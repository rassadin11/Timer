import psutil
from win32gui import GetWindowText, GetForegroundWindow
import time
import eel
from plyer import notification
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

def show_notification(title, message):
    notification.notify(title = title, message = message, app_icon = '../timer.ico', timeout = 5)

eel.init('front')

@eel.expose()

def timer():
    full_time = working_time = 0
    lost_seconds = False
    stop_or_not = True

    while True:
        full_time += 1
        active_window = GetWindowText(GetForegroundWindow())[-18:]
        
        if working_time > 0:
            if working_time % 21600 == 0:
                show_notification('Timer', "You're programming for 6 hours! Great work!")
            elif working_time % 18000 == 0:
                show_notification('Timer', "You're programming for 5 hours! So nice!")
            elif working_time % 14400 == 0:
                show_notification('Timer', "You're programming for 4 hours! Wondeful day for programming!")
            elif working_time % 10800 == 0:
                show_notification('Timer', "You're programming for 3 hours! Good result!")
            elif working_time % 7200 == 0:
                show_notification('Timer', "You're programming for 2 hours! Not bad!")
            elif working_time % 3600 == 0:
                show_notification('Timer', "You're programming for 1 hour! This is only the beginning! Try harder =)")

        if eel.check_start()() == True:
            active_window = 'Visual Studio Code'

            full_time = 0
            stop_or_not = True
        
        if eel.check_stop()() == True:
            lost_seconds = True

        obj = eel.addExcelPosts()()

        def addPost(obj):
            flag = True
                    
            path = 'C:\Python\Timer-master\dist\list.xlsx'

            wb = load_workbook(filename=path)

            if obj['local_variable'] == False:
                return

            try:
                if flag == True:
                    colors = ['6600ff', '7030A0', '002060', '00B0F0', '00B050', '95D050', 'FF0FF0', 'FFC000', 'FF0000', 'C000000']
                    color = ''
                    font_color = '000000'

                    if obj['tm'] > '6:00:00':
                        color = colors[0]
                        font_color = 'ffffff'
                    if (obj['tm'] >= '5:00:00') and (obj['tm'] < '6:00:00'):
                        color = colors[1]
                        font_color = 'ffffff'
                    if (obj['tm'] >= '4:00:00') and (obj['tm'] < '5:00:00'):
                        color = colors[2]
                        font_color = 'ffffff'
                    if (obj['tm'] >= '3:00:00') and (obj['tm'] < '4:00:00'):
                        color = colors[3]
                    if (obj['tm'] >= '2:30:00') and (obj['tm'] < '3:00:00'):
                        color = colors[4]
                    if (obj['tm'] >= '1:30:00') and (obj['tm'] < '2:30:00'):
                        color = colors[5]
                    if (obj['tm'] >= '1:00:00') and (obj['tm'] < '1:30:00'):
                        color = colors[6]
                    if (obj['tm'] >= '0:30:00') and (obj['tm'] < '1:00:00'):
                        color = colors[7]
                    if (obj['tm'] > '0:00:00') and (obj['tm'] < '0:30:00'):
                        color = colors[8]
                    if (obj['tm'] == '0:00:00'):
                        color = colors[9]

                    for cells in wb[wb.sheetnames[0]]['A:K']:
                        for rows in cells:
                            if rows.value != None:
                                if obj['date'] == str(rows.value):
                                    rows.fill = PatternFill("solid", start_color=f"{color}", end_color=f"{color}") # work =)
                                    rows.font = Font(color = f"{font_color}", bold=True)
                                    wb.save(path)

                                    new_coords = f'{get_column_letter(rows.column + 1)}{rows.row}'

                                    wb.save(path)

                                    wb[wb.sheetnames[0]][new_coords] = obj['text']
                                    wb[wb.sheetnames[0]][new_coords].alignment = Alignment(horizontal="center", vertical="center")
                                    wb[wb.sheetnames[0]][new_coords].font = Font(color = f"{font_color}")

                                    wb.save(path)

                                    wb[wb.sheetnames[0]][new_coords].fill = PatternFill("solid", start_color=f"{color}", end_color=f"{color}") # work =)

                                    wb.save(path)

                                    new_coords = f'{get_column_letter(rows.column)}{rows.row + 2}'

                                    wb.save(path)

                                    wb[wb.sheetnames[0]][new_coords] = obj['tm']
                                    wb[wb.sheetnames[0]][new_coords].alignment = Alignment(horizontal="center", vertical="center")
                                    wb[wb.sheetnames[0]][new_coords].font = Font(color = f"{font_color}", bold=True)
                                    
                                    wb.save(path)

                                    wb[wb.sheetnames[0]][new_coords].fill = PatternFill("solid", start_color=f"{color}", end_color=f"{color}") # work =)

                                    wb.save(path)
            except: 
                print('error')

            wb.save(path)

        addPost(obj)

        stop_or_not = eel.stop_or_not()()

        print(stop_or_not)

        try:
            working_time += int(eel.add_time()())
        except:
            continue
        
        if active_window == 'Visual Studio Code':
            working_time += 1
            lost_seconds = False

        else:
            if lost_seconds == False:
                flag = False
                working_time += 1
                if stop_or_not == True:
                    if full_time % 180 == 0:

                        for proc in psutil.process_iter():
                            try:
                                processName = proc.name()

                                if flag == False:
                                    if processName == 'Code.exe':
                                        flag = True
                                        lost_seconds = True
                            
                            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                                pass
        
        eel.show_time(working_time)

        time.sleep(1)

eel.start('index.html', size = (800, 1000))